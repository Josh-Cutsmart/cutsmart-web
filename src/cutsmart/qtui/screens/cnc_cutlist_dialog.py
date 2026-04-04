from __future__ import annotations
import re
from pathlib import Path

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor, QBrush, QFont, QPixmap
from PySide6.QtWidgets import (
    QCheckBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)


class CNCCutlistDialog(QDialog):
    COLUMNS = [
        "ID",
        "Room",
        "Part Type",
        "Board Type",
        "Part Name",
        "Height",
        "Width",
        "Depth",
        "Quantity",
        "Clashing",
        "Grain",
        "Information",
    ]

    def __init__(
        self,
        project_name: str,
        designer_name: str,
        rows: list[dict],
        source_rows: list[dict] | None = None,
        visibility_map: dict[str, bool] | None = None,
        on_visibility_changed=None,
        collapsed_part_types: list[str] | None = None,
        on_collapsed_changed=None,
        show_grain_column: bool = True,
        board_display_map: dict[str, str] | None = None,
        board_sheet_size_map: dict[str, str] | None = None,
        board_edging_map: dict[str, str] | None = None,
        part_type_colors: dict[str, str] | None = None,
        cabinetry_part_types: dict[str, bool] | None = None,
        theme_color: str | None = None,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._project_name = str(project_name or "Project")
        self._designer_name = str(designer_name or "")
        self.setWindowTitle(f"CNC Cutlist - {project_name or 'Project'}")
        self.setModal(False)
        self.setWindowModality(Qt.WindowModality.NonModal)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, False)
        self.setWindowFlag(Qt.WindowType.WindowMaximizeButtonHint, True)
        self.setWindowFlag(Qt.WindowType.WindowMinimizeButtonHint, True)
        self.resize(1480, 900)
        self.setWindowState(self.windowState() | Qt.WindowState.WindowMaximized)
        QTimer.singleShot(0, self.showMaximized)

        self._board_display_map = {str(k or "").strip(): str(v or "").strip() for k, v in (board_display_map or {}).items()}
        self._board_sheet_size_map = {str(k or "").strip(): str(v or "").strip() for k, v in (board_sheet_size_map or {}).items()}
        self._board_edging_map = {str(k or "").strip(): str(v or "").strip() for k, v in (board_edging_map or {}).items()}
        self._part_type_colors: dict[str, str] = {}
        for k, v in (part_type_colors or {}).items():
            key_raw = str(k or "").strip()
            if not key_raw:
                continue
            color_val = self._normalize_color(v)
            if not color_val:
                continue
            self._part_type_colors[key_raw.lower()] = color_val
            self._part_type_colors[self._type_key(key_raw)] = color_val
        self._cabinetry_part_types = {self._type_key(k): bool(v) for k, v in (cabinetry_part_types or {}).items() if str(k or "").strip()}
        qc = QColor(str(theme_color or "").strip())
        self._theme_color = qc.name().upper() if qc.isValid() else "#7FA3B9"
        self._all_rows = [dict(r) for r in (rows or []) if isinstance(r, dict)]
        self._source_rows = [dict(r) for r in (source_rows or []) if isinstance(r, dict)]
        self._source_rows_by_key: dict[str, dict] = {}
        for i, r in enumerate(self._source_rows):
            if not isinstance(r, dict):
                continue
            raw_key = self._row_get(r, "__cutlist_key", "__id")
            key = str(raw_key or "").strip() or str(i)
            self._source_rows_by_key[key] = r
        self._visibility_map = {str(k): bool(v) for k, v in (visibility_map or {}).items()}
        self._on_visibility_changed = on_visibility_changed
        self._on_collapsed_changed = on_collapsed_changed
        self._show_grain_column = bool(show_grain_column)
        self._collapsed_part_types: set[str] = set(
            self._type_key(v) for v in (collapsed_part_types or []) if str(v or "").strip()
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        top = QFrame()
        top.setObjectName("cncTopBar")
        top.setStyleSheet("QFrame#cncTopBar { background:#FFFFFF; border:1px solid #D7DEE8; border-radius:14px; }")
        top.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        top.setMinimumHeight(50)
        top.setMaximumHeight(54)
        top_l = QHBoxLayout(top)
        top_l.setContentsMargins(14, 8, 14, 8)
        top_l.setSpacing(6)
        cnc_icon = QLabel()
        cnc_icon.setStyleSheet("QLabel { background:transparent; border:none; }")
        icon_path = Path(__file__).resolve().parent.parent / "assets" / "icons" / "shapes.png"
        icon_pix = QPixmap(str(icon_path)) if icon_path.exists() else QPixmap()
        if not icon_pix.isNull():
            cnc_icon.setPixmap(icon_pix.scaled(18, 18, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        cnc_icon.setFixedSize(20, 20)
        top_l.addWidget(cnc_icon, 0, Qt.AlignmentFlag.AlignVCenter)
        title = QLabel("CNC CUTLIST")
        title.setStyleSheet("QLabel { color:#0F2A4A; font-size:13px; font-weight:800; letter-spacing:1px; background:transparent; border:none; }")
        top_l.addWidget(title, 0, Qt.AlignmentFlag.AlignVCenter)
        title_div = QLabel("  |  ")
        title_div.setStyleSheet("QLabel { color:#64748B; font-size:13px; font-weight:700; background:transparent; border:none; }")
        top_l.addWidget(title_div, 0, Qt.AlignmentFlag.AlignVCenter)
        job_name = QLabel(str(project_name or "-"))
        job_name.setStyleSheet("QLabel { color:#334155; font-size:13px; font-weight:700; background:transparent; border:none; }")
        top_l.addWidget(job_name, 1, Qt.AlignmentFlag.AlignVCenter)
        export_btn = QPushButton("Export .xlsx")
        export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        export_btn.setStyleSheet(
            "QPushButton { background:#FFFFFF; color:#0F172A; border:1px solid #D4DAE6; border-radius:10px; padding:7px 12px; font-size:12px; font-weight:700; }"
            "QPushButton:hover { background:#F8FAFC; border-color:#B9C4D8; }"
            "QPushButton:pressed { background:#EEF2F7; }"
        )
        export_btn.clicked.connect(self._export_xlsx)
        top_l.addWidget(export_btn, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        root.addWidget(top, 0)

        content = QWidget()
        content_l = QHBoxLayout(content)
        content_l.setContentsMargins(0, 0, 0, 0)
        content_l.setSpacing(10)

        self._boards_scroll = QScrollArea()
        self._boards_scroll.setWidgetResizable(True)
        self._boards_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self._boards_host = QWidget()
        self._boards_host.setStyleSheet("QWidget { background: transparent; border: none; }")
        self._boards_host_l = QVBoxLayout(self._boards_host)
        self._boards_host_l.setContentsMargins(0, 0, 0, 0)
        self._boards_host_l.setSpacing(14)
        self._boards_scroll.setWidget(self._boards_host)
        content_l.addWidget(self._boards_scroll, 1)

        vis_card = QFrame()
        vis_card.setStyleSheet("QFrame { background:#FFFFFF; border:1px solid #E4E6EC; border-radius:10px; }")
        vis_card_l = QVBoxLayout(vis_card)
        vis_card_l.setContentsMargins(10, 10, 10, 10)
        vis_card_l.setSpacing(8)
        vis_head_row = QWidget()
        vis_head_row_l = QHBoxLayout(vis_head_row)
        vis_head_row_l.setContentsMargins(0, 0, 0, 0)
        vis_head_row_l.setSpacing(8)
        vis_title = QLabel("Edit Visibility")
        vis_title.setStyleSheet("QLabel { color:#111827; font-size:13px; font-weight:800; background:transparent; border:none; }")
        vis_head_row_l.addWidget(vis_title, 1, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self._show_all_btn = QPushButton("Show All")
        self._show_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._show_all_btn.setStyleSheet(
            "QPushButton { background:#FFFFFF; color:#0F172A; border:1px solid #D4DAE6; border-radius:8px; padding:4px 10px; font-size:11px; font-weight:700; }"
            "QPushButton:hover { background:#F8FAFC; border-color:#B9C4D8; }"
            "QPushButton:pressed { background:#EEF2F7; }"
        )
        self._show_all_btn.clicked.connect(self._show_all_visibility_rows)
        vis_head_row_l.addWidget(self._show_all_btn, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        vis_card_l.addWidget(vis_head_row)
        self._visibility_search = QLineEdit()
        self._visibility_search.setPlaceholderText("Search pieces...")
        self._visibility_search.setStyleSheet(
            "QLineEdit { background:#FFFFFF; border:1px solid #D4DAE6; border-radius:8px; padding:6px 8px; font-size:12px; color:#111827; }"
            "QLineEdit:focus { border:1px solid #9FB6DA; }"
        )
        self._visibility_search.textChanged.connect(self._apply_visibility_search)
        vis_card_l.addWidget(self._visibility_search)
        self._visibility_scroll = QScrollArea()
        self._visibility_scroll.setWidgetResizable(True)
        self._visibility_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._visibility_scroll.setStyleSheet("QScrollArea { background: #FFFFFF; border: none; }")
        self._visibility_host = QWidget()
        self._visibility_host.setStyleSheet("QWidget { background:#FFFFFF; border:none; }")
        self._visibility_host_l = QVBoxLayout(self._visibility_host)
        self._visibility_host_l.setContentsMargins(0, 0, 0, 0)
        self._visibility_host_l.setSpacing(6)
        self._visibility_scroll.setWidget(self._visibility_host)
        vis_card_l.addWidget(self._visibility_scroll, 1)
        content_l.addWidget(vis_card, 0)
        vis_card.setFixedWidth(360)

        root.addWidget(content, 1)

        self._rebuild_table()
        self._build_visibility_panel()

    @staticmethod
    def _row_get(row: dict, *keys: str) -> str:
        if not isinstance(row, dict):
            return ""
        for key in keys:
            value = row.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return ""

    def _board_label(self, raw_board: str) -> str:
        key = str(raw_board or "").strip()
        return self._board_display_map.get(key) or key or "Unknown Board"

    def _board_base_label(self, raw_board: str) -> str:
        label = self._board_label(raw_board)
        return re.sub(r"^\[[^\]]+\]\s*", "", str(label or "").strip()).strip() or str(label or "").strip() or "Unknown Board"

    def _board_header_label(self, raw_board: str) -> str:
        key = str(raw_board or "").strip()
        base = self._board_base_label(key)
        full_sheet = str(self._board_sheet_size_map.get(key) or "").strip()
        edging = str(self._board_edging_map.get(key) or "").strip()
        sep = "   |   "
        if full_sheet:
            if edging:
                return f"{full_sheet}{sep}{base}{sep}Edging: {edging}"
            return f"{full_sheet}{sep}{base}"
        if edging:
            return f"{base}{sep}Edging: {edging}"
        return base

    def _is_dark(self, hex_color: str) -> bool:
        c = QColor(str(hex_color or "#FFFFFF"))
        if not c.isValid():
            c = QColor("#FFFFFF")
        luminance = (0.299 * c.red() + 0.587 * c.green() + 0.114 * c.blue())
        return luminance < 145

    def _lighten_color(self, hex_color: str, ratio: float = 0.78) -> str:
        base = QColor(str(hex_color or "#E8EEF7"))
        if not base.isValid():
            base = QColor("#E8EEF7")
        r = max(0.0, min(1.0, float(ratio)))
        rr = int(base.red() + (255 - base.red()) * r)
        gg = int(base.green() + (255 - base.green()) * r)
        bb = int(base.blue() + (255 - base.blue()) * r)
        return f"#{rr:02X}{gg:02X}{bb:02X}"

    def _normalize_color(self, value) -> str:
        txt = str(value or "").strip()
        if not txt:
            return ""
        c = QColor(txt)
        if c.isValid():
            return c.name().upper()
        m = re.search(r"#(?:[0-9A-Fa-f]{6}|[0-9A-Fa-f]{3})", txt)
        if m:
            c2 = QColor(m.group(0))
            if c2.isValid():
                return c2.name().upper()
        return ""

    def _mk_item(self, text: str, align: Qt.AlignmentFlag | None = None) -> QTableWidgetItem:
        item = QTableWidgetItem(str(text or ""))
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        if align is not None:
            item.setTextAlignment(int(align))
        return item

    def _format_clashing_display(self, value: str) -> str:
        txt = str(value or "").strip()
        if not txt:
            return ""
        parts = [p for p in txt.split() if p]
        if len(parts) <= 1:
            return txt
        return "   ".join(parts)

    def _boolish(self, value) -> bool:
        if isinstance(value, bool):
            return value
        txt = str(value or "").strip().lower()
        if txt in ("0", "false", "no", "off", "n"):
            return False
        if txt in ("1", "true", "yes", "on", "y"):
            return True
        return bool(value)

    def _is_row_visible_for_nesting(self, row: dict) -> bool:
        idx = str(row.get("__cutlist_key") or "").strip()
        if idx:
            return bool(self._visibility_map.get(idx, True))
        return self._boolish(row.get("includeInNesting", True))

    def _rebuild_table(self) -> None:
        filtered = [dict(r) for r in self._all_rows if self._is_row_visible_for_nesting(r)]
        self._build_rows(filtered)

    def _normalized_rows(self, rows: list[dict]) -> list[dict]:
        normalized: list[dict] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            board_raw = self._row_get(row, "board", "boardType", "material")
            name = self._row_get(row, "name", "partName")
            part_type = self._row_get(row, "partType", "part_type", "type")
            qty = self._row_get(row, "quantity", "qty")
            if not name or not part_type or not board_raw or not qty:
                continue
            normalized.append(
                {
                    "board_raw": board_raw,
                    "board": self._board_base_label(board_raw),
                    "board_header": self._board_header_label(board_raw),
                    "job_section": self._row_get(row, "room", "section", "jobSection"),
                    "part_type": part_type,
                    "name": name,
                    "height": self._row_get(row, "height", "h"),
                    "width": self._row_get(row, "width", "w"),
                    "depth": self._row_get(row, "depth", "d"),
                    "qty": qty,
                    "clashing": self._row_get(row, "clashing"),
                    "grain": self._row_get(row, "grain"),
                    "info": self._row_get(row, "information", "info", "notes"),
                    "adjustable_shelves": self._row_get(row, "adjustableShelf", "adjustableShelves", "adjustable_shelf"),
                    "fixed_shelves": self._row_get(row, "fixedShelf", "fixedShelves", "fixed_shelf"),
                    "adjustable_drilling": self._row_get(
                        row,
                        "adjustableShelfDrilling",
                        "adjustable_drilling",
                        "adjustableDrilling",
                        "adjustable_shelf_drilling",
                    ),
                    "fixed_drilling": self._row_get(
                        row,
                        "fixedShelfDrilling",
                        "fixed_drilling",
                        "fixedDrilling",
                        "fixed_shelf_drilling",
                    ),
                }
            )
        normalized.sort(key=self._cnc_sort_key)
        return normalized

    @staticmethod
    def _piece_kind_sort_rank(name: str) -> int:
        txt = str(name or "").strip().lower()
        if re.search(r"\bbottom\b", txt):
            return 0
        if re.search(r"\bback\b", txt):
            return 1
        return 2

    def _cnc_sort_key(self, row: dict) -> tuple[str, str, int, str]:
        return (
            str(row.get("board_header") or row.get("board") or "").lower(),
            str(row.get("part_type") or "").lower(),
            int(self._piece_kind_sort_rank(str(row.get("name") or ""))),
            str(row.get("name") or "").lower(),
        )

    def _is_cabinetry_type(self, part_type: str) -> bool:
        return bool(self._cabinetry_part_types.get(self._type_key(part_type), False))

    def _parse_cabinet_shelves_summary(self, value: object) -> dict[str, str]:
        out = {
            "fixed_qty": "",
            "fixed_drilling": "",
            "adjustable_qty": "",
            "adjustable_drilling": "",
        }
        txt = str(value or "").strip()
        if not txt:
            return out
        for line in re.split(r"[\r\n]+", txt):
            token = str(line or "").strip()
            if not token:
                continue
            m = re.match(r"(?i)^\s*(fixed|adjustable)(?:\s*shelf)?\s*:\s*(.+?)\s*$", token)
            if not m:
                continue
            kind = str(m.group(1) or "").strip().lower()
            rest = str(m.group(2) or "").strip()
            qty = ""
            drill = ""
            if "|" in rest:
                left, right = rest.split("|", 1)
                qty = str(left or "").strip()
                drill = str(right or "").strip()
            else:
                qty = rest
                drill = ""
            drill = re.sub(r"(?i)^\s*drilling\s*:\s*", "", drill).strip()
            if kind == "fixed":
                out["fixed_qty"] = qty
                out["fixed_drilling"] = drill
            elif kind == "adjustable":
                out["adjustable_qty"] = qty
                out["adjustable_drilling"] = drill
        return out

    @staticmethod
    def _first_nonempty(*values: object) -> str:
        for v in values:
            txt = str(v or "").strip()
            if txt:
                return txt
        return ""

    def _find_source_row_for(self, row: dict) -> dict | None:
        if not isinstance(row, dict):
            return None
        raw_embedded = row.get("__raw_row")
        if isinstance(raw_embedded, dict):
            return raw_embedded
        raw_key = self._row_get(row, "__cutlist_key", "__id")
        key = str(raw_key or "").strip()
        if key and key in self._source_rows_by_key:
            return self._source_rows_by_key.get(key)
        rid = str(self._row_get(row, "__id") or "").strip()
        if rid and rid in self._source_rows_by_key:
            return self._source_rows_by_key.get(rid)
        return None

    def _export_xlsx(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except Exception:
            QMessageBox.warning(self, "Export .xlsx", "Export requires openpyxl. Please install openpyxl first.")
            return

        default_name = f"{self._project_name or 'cnc_cutlist'}_cnc_cutlist.xlsx".replace(" ", "_")
        path, _ = QFileDialog.getSaveFileName(self, "Export CNC Cutlist", default_name, "Excel Workbook (*.xlsx)")
        if not path:
            return
        if not str(path).lower().endswith(".xlsx"):
            path = f"{path}.xlsx"

        filtered = [dict(r) for r in self._all_rows if self._is_row_visible_for_nesting(r)]
        normalized = self._normalized_rows(filtered)
        grouped: dict[str, list[dict]] = {}
        cabinetry_rows: list[dict] = []
        for row in normalized:
            if self._is_cabinetry_type(str(row.get("part_type") or "")):
                cabinetry_rows.append(row)
                continue
            board = str(row.get("board_header") or row.get("board") or "Unknown Board")
            grouped.setdefault(board, []).append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = "CNC Cutlist"

        show_grain = bool(self._show_grain_column)
        export_fields: list[tuple[str, int, str]] = [
            ("ID", 1, "center"),
            ("Room", 3, "left"),
            ("Part Type", 2, "left"),
            ("Board Type", 5, "left"),
            ("Part Name", 5, "left"),
            ("Height", 2, "center"),
            ("Width", 2, "center"),
            ("Depth", 2, "center"),
            ("Quantity", 2, "center"),
            ("Clashing", 2, "center"),
        ]
        if show_grain:
            export_fields.append(("Grain", 1, "center"))
            export_fields.append(("Information", 7, "left"))
        else:
            export_fields.append(("Information", 7, "left"))
        total_export_cols = int(sum(int(span) for _name, span, _align in export_fields))

        board_fill = PatternFill("solid", fgColor="111111")
        header_fill = PatternFill("solid", fgColor=str(self._theme_color).replace("#", ""))
        white_font = Font(color="FFFFFF", bold=True)
        dark_font_bold = Font(color=("FFFFFF" if self._is_dark(self._theme_color) else "0F172A"), bold=True)
        body_font = Font(color="111827")
        zebra_fill = PatternFill("solid", fgColor="F5F5F5")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        no_side = Side(style=None)
        thin = Side(style="thin", color="000000")
        bottom_strong = Side(style="medium", color="000000")
        part_type_fill_cache: dict[str, PatternFill] = {}
        part_type_font_cache: dict[str, Font] = {}
        field_end_cols: list[int] = []
        _run_end = 0
        for _nm, _sp, _al in export_fields:
            _run_end += int(_sp)
            field_end_cols.append(_run_end)

        row_ptr = 1
        part_counter = 1
        export_data_row_h = 15
        for board_name in sorted(grouped.keys(), key=lambda x: x.lower()):
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=total_export_cols)
            c = ws.cell(row=row_ptr, column=1, value=board_name)
            c.fill = board_fill
            c.font = white_font
            c.alignment = left
            ws.row_dimensions[row_ptr].height = 22
            row_ptr += 1

            centered_header_labels = {"Room", "Part Type", "Board Type", "Part Name", "Information"}
            col_run = 1
            for label, span, align_mode in export_fields:
                c1 = col_run
                c2 = col_run + int(span) - 1
                if c2 > c1:
                    ws.merge_cells(start_row=row_ptr, start_column=c1, end_row=row_ptr, end_column=c2)
                for cc in range(c1, c2 + 1):
                    h = ws.cell(row=row_ptr, column=cc)
                    if cc == c1:
                        h.value = label
                    h.fill = header_fill
                    h.font = dark_font_bold
                    h.alignment = center if label in centered_header_labels else (center if align_mode == "center" else left)
                    h.border = Border(
                        left=thin if (cc == c1 and c1 == 1) else no_side,
                        right=thin if cc == c2 else no_side,
                        top=thin,
                        bottom=thin,
                    )
                col_run = c2 + 1
            ws.row_dimensions[row_ptr].height = 20
            row_ptr += 1

            zebra = False
            board_data_start = row_ptr
            for row in grouped.get(board_name, []):
                values = {
                    "ID": str(part_counter),
                    "Room": str(row.get("job_section") or ""),
                    "Part Type": str(row.get("part_type") or ""),
                    "Board Type": str(row.get("board") or ""),
                    "Part Name": str(row.get("name") or ""),
                    "Height": str(row.get("height") or ""),
                    "Width": str(row.get("width") or ""),
                    "Depth": str(row.get("depth") or ""),
                    "Quantity": str(row.get("qty") or ""),
                    "Clashing": self._format_clashing_display(str(row.get("clashing") or "")),
                    "Grain": str(row.get("grain") or ""),
                    "Information": str(row.get("info") or ""),
                }
                col_run = 1
                for label, span, align_mode in export_fields:
                    c1 = col_run
                    c2 = col_run + int(span) - 1
                    if c2 > c1:
                        ws.merge_cells(start_row=row_ptr, start_column=c1, end_row=row_ptr, end_column=c2)
                    txt = values.get(label, "")
                    if label == "Part Type":
                        part_key = self._type_key(str(values.get("Part Type") or ""))
                        part_hex = (
                            self._part_type_colors.get(str(values.get("Part Type") or "").strip().lower())
                            or self._part_type_colors.get(part_key)
                            or self._part_type_colors.get(str(values.get("Part Type") or "").strip())
                            or ""
                        )
                        part_hex = self._normalize_color(part_hex) if part_hex else ""
                    else:
                        part_hex = ""
                    for cc in range(c1, c2 + 1):
                        cell = ws.cell(row=row_ptr, column=cc)
                        if cc == c1:
                            cell.value = txt
                        cell.font = body_font
                        cell.alignment = center if align_mode == "center" else left
                        if label == "Part Type" and part_hex:
                            fill = part_type_fill_cache.get(part_hex)
                            if fill is None:
                                fill = PatternFill("solid", fgColor=part_hex.replace("#", ""))
                                part_type_fill_cache[part_hex] = fill
                            pfont = part_type_font_cache.get(part_hex)
                            if pfont is None:
                                pfont = Font(color=("FFFFFF" if self._is_dark(part_hex) else "0F172A"), bold=False)
                                part_type_font_cache[part_hex] = pfont
                            cell.fill = fill
                            cell.font = pfont
                        elif zebra:
                            cell.fill = zebra_fill
                        cell.border = Border(
                            left=thin if (cc == c1 and c1 == 1) else no_side,
                            right=thin if cc == c2 else no_side,
                            top=no_side,
                            bottom=no_side,
                        )
                    col_run = c2 + 1
                part_counter += 1
                zebra = not zebra
                ws.row_dimensions[row_ptr].height = export_data_row_h
                row_ptr += 1
            board_data_end = row_ptr - 1
            if board_data_end >= board_data_start:
                for col_idx in range(1, total_export_cols + 1):
                    c2 = ws.cell(row=board_data_end, column=col_idx)
                    c2.border = Border(
                        left=thin if col_idx == 1 else no_side,
                        right=thin if col_idx in field_end_cols else no_side,
                        top=no_side,
                        bottom=bottom_strong,
                    )
            row_ptr += 1

        if cabinetry_rows:
            row_ptr += 1
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=total_export_cols)
            ch = ws.cell(row=row_ptr, column=1, value="Cabinets")
            ch.fill = board_fill
            ch.font = white_font
            ch.alignment = left
            ws.row_dimensions[row_ptr].height = 22
            row_ptr += 1

            cab_side = Side(style="thin", color="000000")
            cab_no_border = Border(left=no_side, right=no_side, top=no_side, bottom=no_side)
            cab_label_font = Font(color="000000", bold=True)
            cab_value_font = Font(color="000000", bold=False)
            cab_header_text = "FFFFFF" if self._is_dark(self._theme_color) else "0F172A"
            cab_header_label_font = Font(color=cab_header_text, bold=True)
            cab_header_value_font = Font(color=cab_header_text, bold=False)
            cab_alt_fill = PatternFill("solid", fgColor="F5F5F5")
            cab_white_fill = PatternFill("solid", fgColor="FFFFFF")
            cab_left_fill = PatternFill("solid", fgColor="D9DDE4")

            def _drilling_value(value) -> str:
                txt = str(value or "").strip().lower()
                if not txt:
                    return "No"
                if txt in ("even spacing", "even", "spacing", "equal spacing", "evenly spaced", "even-spaced"):
                    return "Even Spacing"
                if txt in ("centre", "center", "centred", "centered"):
                    return "Centre"
                if txt in ("no", "no drilling", "none", "off", "false", "0"):
                    return "No"
                return str(value or "No")

            def _set_cell(rr: int, cc: int, text: str, *, bold: bool = False, align_left: bool = True, alt: bool = False) -> None:
                c = ws.cell(row=rr, column=cc, value=str(text or ""))
                c.font = cab_label_font if bold else cab_value_font
                c.alignment = left if align_left else center
                c.fill = cab_alt_fill if alt else cab_white_fill
                c.border = cab_no_border

            def _merge_write(
                rr: int,
                c1: int,
                c2: int,
                text: str,
                *,
                bold: bool = False,
                alt: bool = False,
                align_mode: str = "left",
                fill_override=None,
                font_override=None,
            ) -> None:
                c1 = int(c1)
                c2 = int(max(c1, c2))
                if c2 > c1:
                    ws.merge_cells(start_row=rr, start_column=c1, end_row=rr, end_column=c2)
                mode = str(align_mode).lower()
                if mode == "right":
                    _align = right
                elif mode == "center":
                    _align = center
                else:
                    _align = left
                for cc in range(c1, c2 + 1):
                    c = ws.cell(row=rr, column=cc)
                    if cc == c1:
                        c.value = str(text or "")
                    c.font = font_override if font_override is not None else (cab_label_font if bold else cab_value_font)
                    c.alignment = _align
                    c.fill = fill_override if fill_override is not None else (cab_alt_fill if alt else cab_white_fill)
                    c.border = cab_no_border

            def _merge_write_segments(rr: int, c1: int, c2: int, segments: list[tuple[str, bool]], *, alt: bool = False) -> None:
                c1 = int(c1)
                c2 = int(max(c1, c2))
                width = c2 - c1 + 1
                if width <= 0:
                    return
                clean_segments = [(str(t or ""), bool(b)) for t, b in (segments or []) if str(t or "")]
                if not clean_segments:
                    _merge_write(rr, c1, c2, "", bold=False, alt=alt)
                    return
                if len(clean_segments) == 1:
                    _merge_write(rr, c1, c2, clean_segments[0][0], bold=clean_segments[0][1], alt=alt)
                    return
                weights = [max(1, len(t.strip()) if t.strip() else 1) for t, _ in clean_segments]
                total_w = sum(weights) or 1
                spans = [max(1, int(round((w / total_w) * width))) for w in weights]
                # normalize to exact width
                while sum(spans) > width:
                    i = max(range(len(spans)), key=lambda ix: spans[ix])
                    if spans[i] > 1:
                        spans[i] -= 1
                    else:
                        break
                while sum(spans) < width:
                    i = max(range(len(weights)), key=lambda ix: weights[ix])
                    spans[i] += 1
                col = c1
                for idx, ((txt, is_bold), span) in enumerate(zip(clean_segments, spans)):
                    end_col = min(c2, col + int(span) - 1)
                    if idx == len(clean_segments) - 1:
                        end_col = c2
                    seg_txt = str(txt or "").strip().lower()
                    _merge_write(
                        rr,
                        col,
                        end_col,
                        txt,
                        bold=is_bold,
                        alt=alt,
                        align_mode=(
                            "right"
                            if seg_txt in {"h:", "w:", "d:"}
                            else ("center" if seg_txt == "x" else "left")
                        ),
                    )
                    col = end_col + 1
                    if col > c2:
                        break

            # Split cabinetry into 3 blocks per row, matching the window's card grid feel.
            total_cols = max(1, int(total_export_cols))
            base_w = total_cols // 3
            rem_w = total_cols % 3
            # Give remainder columns to the right-most blocks first so the 3rd cabinet
            # column has enough width for full H/W/D display.
            block_widths = [base_w + (1 if i >= (3 - rem_w) else 0) for i in range(3)]
            block_starts: list[int] = []
            run_col = 1
            for wv in block_widths:
                block_starts.append(run_col)
                run_col += max(1, int(wv))

            cab_entries: list[dict] = []
            has_any_cabinet_info = False
            for row in cabinetry_rows:
                source_row = self._find_source_row_for(row) or {}
                info_raw = str(
                    self._first_nonempty(
                        row.get("info"),
                        self._row_get(source_row, "information", "info", "notes"),
                    )
                    or ""
                )
                info_lines = [ln.strip() for ln in info_raw.replace("\r", "\n").split("\n") if ln.strip()]
                if info_lines:
                    has_any_cabinet_info = True
                cab_entries.append({"row": row, "source_row": source_row, "info_lines": info_lines})

            def _card_total_rows(entry: dict) -> int:
                info_lines = list(entry.get("info_lines") or [])
                info_count = max(1, len(info_lines)) if has_any_cabinet_info else 0
                return 8 + info_count

            num_block_rows = (len(cab_entries) + 2) // 3
            block_row_heights: list[int] = []
            for br in range(num_block_rows):
                chunk = cab_entries[br * 3 : (br * 3) + 3]
                block_row_heights.append(max((_card_total_rows(ent) for ent in chunk), default=0))
            block_row_starts: list[int] = []
            rr = row_ptr
            for hgt in block_row_heights:
                block_row_starts.append(rr)
                rr += hgt

            for idx, entry in enumerate(cab_entries):
                row = dict(entry.get("row") or {})
                source_row = dict(entry.get("source_row") or {})
                info_lines = list(entry.get("info_lines") or [])
                block_col = idx % 3
                block_row = idx // 3
                start_col = block_starts[block_col]
                block_w = max(1, block_widths[block_col])
                r0 = block_row_starts[block_row]
                card_h = _card_total_rows(entry)
                block_h = block_row_heights[block_row]

                name = str(row.get("name") or "-")
                material = str(row.get("board") or "-")
                h = str(row.get("height") or "-")
                w = str(row.get("width") or "-")
                d = str(row.get("depth") or "-")
                qty_txt = str(row.get("qty") or "-")
                size_txt = f"H:{h} x W:{w} x D:{d}"
                board_key = str(
                    self._first_nonempty(
                        self._row_get(source_row, "board", "boardType", "material"),
                        self._row_get(row, "board_raw", "board"),
                    )
                ).strip()
                clashing_txt = str(self._board_edging_map.get(board_key) or "").strip() or "-"

                shelves_summary_row = self._parse_cabinet_shelves_summary(self._row_get(row, "clashing", "shelves"))
                shelves_summary_src = self._parse_cabinet_shelves_summary(self._row_get(source_row, "clashing", "shelves"))

                adj_num = self._first_nonempty(
                    shelves_summary_src.get("adjustable_qty"),
                    shelves_summary_row.get("adjustable_qty"),
                    self._row_get(source_row, "adjustableShelf", "adjustable_shelves", "adjustableShelves"),
                    self._row_get(row, "adjustable_shelves", "adjustableShelf", "adjustableShelves"),
                ) or "-"
                adj_drill = _drilling_value(
                    self._first_nonempty(
                        shelves_summary_src.get("adjustable_drilling"),
                        shelves_summary_row.get("adjustable_drilling"),
                        self._row_get(source_row, "adjustableShelfDrilling", "adjustable_drilling", "adjustableDrilling"),
                        self._row_get(row, "adjustable_drilling", "adjustableShelfDrilling", "adjustableDrilling"),
                    )
                )
                fix_num = self._first_nonempty(
                    shelves_summary_src.get("fixed_qty"),
                    shelves_summary_row.get("fixed_qty"),
                    self._row_get(source_row, "fixedShelf", "fixed_shelves", "fixedShelves"),
                    self._row_get(row, "fixed_shelves", "fixedShelf", "fixedShelves"),
                ) or "-"
                fix_drill = _drilling_value(
                    self._first_nonempty(
                        shelves_summary_src.get("fixed_drilling"),
                        shelves_summary_row.get("fixed_drilling"),
                        self._row_get(source_row, "fixedShelfDrilling", "fixed_drilling", "fixedDrilling"),
                        self._row_get(row, "fixed_drilling", "fixedShelfDrilling", "fixedDrilling"),
                    )
                )

                cab_id = str(part_counter)
                part_counter += 1

                # Row 1: ID + Name on same row.
                # Row 2: Cabinet Size
                # Row 3: Quantity
                # Row 4: Material
                # Row 5: Clashing
                # Row 6: blank spacer
                # Row 7: Adjustable Shelf + Drilling
                # Row 8: Fixed Shelf + Drilling
                end_col = start_col + block_w - 1
                if block_w < 4:
                    _merge_write(r0 + 0, start_col, end_col, f"ID: {cab_id}    Name: {name}", bold=False, alt=False)
                    _merge_write(r0 + 1, start_col, end_col, f"Cabinet Size {size_txt}", bold=False, alt=True)
                    _merge_write(r0 + 2, start_col, end_col, f"Quantity {qty_txt}", bold=False, alt=False)
                    _merge_write(r0 + 3, start_col, end_col, f"Material {material}", bold=False, alt=True)
                    _merge_write(r0 + 4, start_col, end_col, f"Clashing {clashing_txt}", bold=False, alt=False)
                    _merge_write(r0 + 5, start_col, end_col, "", bold=False, alt=True)
                    _merge_write(r0 + 6, start_col, end_col, f"Adjustable Shelf {adj_num}    Drilling: {adj_drill}", bold=False, alt=False)
                    _merge_write(r0 + 7, start_col, end_col, f"Fixed Shelf {fix_num}    Drilling: {fix_drill}", bold=False, alt=True)
                    if has_any_cabinet_info:
                        first_line = f"- {info_lines[0]}" if info_lines else "-"
                        _merge_write(r0 + 8, start_col, end_col, f"Information {first_line}", bold=True, alt=False)
                        for ii, ln in enumerate(info_lines[1:]):
                            _merge_write(r0 + 9 + ii, start_col, end_col, f"- {ln}", bold=False, alt=(ii % 2 == 0))
                else:
                    # Cabinet layout tuning:
                    # - ID title/value kept to 1 column each
                    # - Name given more space
                    # - All row titles use 2-column spans
                    label_w = 2
                    id_label_c1 = start_col
                    id_label_c2 = start_col
                    id_val_col = min(id_label_c2 + 1, end_col)
                    name_label_c1 = min(id_val_col + 2, end_col)
                    # Keep "Name:" to a single column so the cabinet name starts earlier.
                    name_label_c2 = name_label_c1
                    name_val_c1 = min(name_label_c2 + 1, end_col)

                    _merge_write(
                        r0 + 0, id_label_c1, id_label_c2, "ID:", bold=True, alt=False,
                        fill_override=header_fill, font_override=cab_header_label_font,
                    )
                    _merge_write(
                        r0 + 0, id_val_col, id_val_col, cab_id, bold=False, alt=False,
                        fill_override=header_fill, font_override=cab_header_value_font,
                    )
                    if name_label_c1 > (id_val_col + 1):
                        # Preserve full-row fill where Name is shifted right.
                        _merge_write(
                            r0 + 0, id_val_col + 1, name_label_c1 - 1, "", bold=False, alt=False,
                            fill_override=header_fill, font_override=cab_header_value_font,
                        )
                    _merge_write(
                        r0 + 0, name_label_c1, name_label_c2, "Name:", bold=True, alt=False, align_mode="right",
                        fill_override=header_fill, font_override=cab_header_label_font,
                    )
                    _merge_write(
                        r0 + 0, name_val_c1, end_col, name, bold=False, alt=False,
                        fill_override=header_fill, font_override=cab_header_value_font,
                    )

                    left_label_c1 = start_col
                    left_label_c2 = min(start_col + 2, end_col)
                    left_value_c1 = min(left_label_c2 + 1, end_col)

                    _merge_write(
                        r0 + 1, left_label_c1, left_label_c2, "Cabinet Size", bold=True, alt=True,
                        fill_override=cab_left_fill,
                    )
                    size_value_c1 = left_value_c1
                    _merge_write_segments(
                        r0 + 1,
                        size_value_c1,
                        end_col,
                        [
                            ("H:", True),
                            (str(h or "-"), False),
                            (" x ", False),
                            ("W:", True),
                            (str(w or "-"), False),
                            (" x ", False),
                            ("D:", True),
                            (str(d or "-"), False),
                        ],
                        alt=True,
                    )
                    _merge_write(
                        r0 + 2, left_label_c1, left_label_c2, "Quantity", bold=True, alt=False,
                        fill_override=cab_left_fill,
                    )
                    _merge_write(r0 + 2, left_value_c1, end_col, qty_txt, bold=False, alt=False)
                    _merge_write(
                        r0 + 3, left_label_c1, left_label_c2, "Material", bold=True, alt=True,
                        fill_override=cab_left_fill,
                    )
                    _merge_write(r0 + 3, left_value_c1, end_col, material, bold=False, alt=True)
                    _merge_write(
                        r0 + 4, left_label_c1, left_label_c2, "Clashing", bold=True, alt=False,
                        fill_override=cab_left_fill,
                    )
                    _merge_write(r0 + 4, left_value_c1, end_col, clashing_txt, bold=False, alt=False)
                    _merge_write(r0 + 5, left_label_c1, left_label_c2, "", bold=False, alt=True, fill_override=cab_left_fill)
                    if left_value_c1 <= end_col:
                        _merge_write(r0 + 5, left_value_c1, end_col, "", bold=False, alt=True)

                    shelf_title_c1 = left_label_c1
                    shelf_title_c2 = left_label_c2
                    shelf_qty_col = left_value_c1
                    drill_label_c1 = min(shelf_qty_col + 1, end_col)
                    drill_label_c2 = min(drill_label_c1 + (label_w - 1), end_col)
                    drill_value_c1 = min(drill_label_c2 + 1, end_col)

                    _merge_write(
                        r0 + 6, shelf_title_c1, shelf_title_c2, "Adjustable Shelf", bold=True, alt=False,
                        fill_override=cab_left_fill,
                    )
                    _merge_write(r0 + 6, shelf_qty_col, shelf_qty_col, adj_num, bold=False, alt=False)
                    _merge_write(r0 + 6, drill_label_c1, drill_label_c2, "Drilling:", bold=True, alt=False)
                    _merge_write(r0 + 6, drill_value_c1, end_col, adj_drill, bold=False, alt=False)

                    _merge_write(
                        r0 + 7, shelf_title_c1, shelf_title_c2, "Fixed Shelf", bold=True, alt=True,
                        fill_override=cab_left_fill,
                    )
                    _merge_write(r0 + 7, shelf_qty_col, shelf_qty_col, fix_num, bold=False, alt=True)
                    _merge_write(r0 + 7, drill_label_c1, drill_label_c2, "Drilling:", bold=True, alt=True)
                    _merge_write(r0 + 7, drill_value_c1, end_col, fix_drill, bold=False, alt=True)
                    if has_any_cabinet_info:
                        info_r0 = r0 + 8
                        first_line = f"- {info_lines[0]}" if info_lines else "-"
                        _merge_write(
                            info_r0, left_label_c1, left_label_c2, "Information",
                            bold=True, alt=False, fill_override=cab_left_fill,
                        )
                        _merge_write(info_r0, left_value_c1, end_col, first_line, bold=False, alt=False)
                        for ii, ln in enumerate(info_lines[1:]):
                            rr_info = info_r0 + 1 + ii
                            alt_row = (ii % 2 == 0)
                            _merge_write(rr_info, left_label_c1, left_label_c2, "", bold=False, alt=alt_row, fill_override=cab_left_fill)
                            _merge_write(rr_info, left_value_c1, end_col, f"- {ln}", bold=False, alt=alt_row)

                for rr_h in range(r0, r0 + block_h):
                    ws.row_dimensions[rr_h].height = export_data_row_h
                # Cabinet box border: outer left/right and bottom only.
                for rr_b in range(r0, r0 + block_h):
                    for cc in range(start_col, end_col + 1):
                        cedge = ws.cell(row=rr_b, column=cc)
                        cedge.border = Border(
                            left=cab_side if cc == start_col else no_side,
                            right=cab_side if cc == end_col else no_side,
                            top=no_side,
                            bottom=cab_side if rr_b == (r0 + block_h - 1) else no_side,
                        )

            rows_used = sum(block_row_heights)
            row_ptr += rows_used + 1

        from openpyxl.utils import get_column_letter
        for col_idx in range(1, total_export_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 7.0

        try:
            wb.save(path)
            QMessageBox.information(self, "Export .xlsx", f"CNC cutlist exported:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export .xlsx", str(exc))

    def _visibility_label(self, row: dict) -> str:
        part_name = self._row_get(row, "name", "partName") or "Piece"
        room = self._row_get(row, "room", "section", "jobSection") or "-"
        return f"{part_name} {room}".strip()

    def _set_visibility(self, cutlist_index: str, checked: bool) -> None:
        key = str(cutlist_index or "").strip()
        if not key:
            return
        self._visibility_map[key] = bool(checked)
        if callable(self._on_visibility_changed):
            try:
                self._on_visibility_changed(dict(self._visibility_map))
            except Exception:
                pass
        self._rebuild_table()
        self._refresh_group_header_checks()

    def _show_all_visibility_rows(self) -> None:
        changed = False
        for row in self._source_rows:
            idx = str(row.get("__cutlist_key") or "").strip()
            if not idx:
                continue
            if not bool(self._visibility_map.get(idx, True)):
                self._visibility_map[idx] = True
                changed = True
        if changed and callable(self._on_visibility_changed):
            try:
                self._on_visibility_changed(dict(self._visibility_map))
            except Exception:
                pass
        self._rebuild_table()
        self._build_visibility_panel()

    @staticmethod
    def _type_key(value: str) -> str:
        return " ".join(str(value or "").strip().lower().split())

    def _set_part_type_visibility(self, part_type_key: str, checked: bool) -> None:
        part_key = self._type_key(part_type_key)
        target_visible = bool(checked)
        changed = False
        for row in self._source_rows:
            idx = str(row.get("__cutlist_key") or "").strip()
            if not idx:
                continue
            row_key = self._type_key(self._row_get(row, "partType", "part_type", "type") or "Unassigned")
            if row_key != part_key:
                continue
            if bool(self._visibility_map.get(idx, True)) != bool(target_visible):
                self._visibility_map[idx] = bool(target_visible)
                changed = True
        for _row_widget, _blob, group_key, row_cb in (getattr(self, "_visibility_row_widgets", None) or []):
            if self._type_key(group_key) != part_key:
                continue
            if isinstance(row_cb, QCheckBox):
                row_cb.blockSignals(True)
                row_cb.setChecked(bool(target_visible))
                row_cb.blockSignals(False)
        if changed and callable(self._on_visibility_changed):
            try:
                self._on_visibility_changed(dict(self._visibility_map))
            except Exception:
                pass
        self._rebuild_table()
        self._refresh_group_header_checks()

    def _refresh_group_header_checks(self) -> None:
        for header_widget, group_name, cb, _collapse_btn, _group_key in (getattr(self, "_visibility_group_widgets", None) or []):
            _ = header_widget
            if not isinstance(cb, QCheckBox):
                continue
            group_key = self._type_key(group_name)
            states: list[bool] = []
            for row in self._source_rows:
                idx = str(row.get("__cutlist_key") or "").strip()
                if not idx:
                    continue
                row_key = self._type_key(self._row_get(row, "partType", "part_type", "type") or "Unassigned")
                if row_key != group_key:
                    continue
                states.append(bool(self._visibility_map.get(idx, self._boolish(row.get("includeInNesting", True)))))
            all_checked = bool(states) and all(states)
            cb.blockSignals(True)
            cb.setChecked(all_checked)
            cb.blockSignals(False)

    def _toggle_part_type_collapsed(self, group_key: str, button: QPushButton) -> None:
        key = self._type_key(group_key)
        if not key:
            return
        if key in self._collapsed_part_types:
            self._collapsed_part_types.discard(key)
        else:
            self._collapsed_part_types.add(key)
        if isinstance(button, QPushButton):
            button.setText("+" if key in self._collapsed_part_types else "-")
        if callable(self._on_collapsed_changed):
            try:
                self._on_collapsed_changed(sorted(self._collapsed_part_types))
            except Exception:
                pass
        self._apply_visibility_search()

    def _clear_layout(self, layout: QVBoxLayout) -> None:
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

    def _build_visibility_panel(self) -> None:
        if not isinstance(getattr(self, "_visibility_host_l", None), QVBoxLayout):
            return
        self._clear_layout(self._visibility_host_l)
        self._visibility_group_widgets: list[tuple[QWidget, str, QCheckBox, QPushButton, str]] = []
        self._visibility_row_widgets: list[tuple[QWidget, str, str, QCheckBox]] = []
        rows = [dict(r) for r in self._source_rows if str(r.get("__cutlist_key") or "").strip()]
        rows.sort(key=lambda r: (self._row_get(r, "partType", "part_type", "type").lower(), self._row_get(r, "name", "partName").lower()))
        grouped: dict[str, list[dict]] = {}
        for row in rows:
            ptype = self._row_get(row, "partType", "part_type", "type") or "Unassigned"
            grouped.setdefault(ptype, []).append(row)
        for part_type in sorted(grouped.keys(), key=lambda x: x.lower()):
            pkey = self._type_key(part_type)
            raw_color = str(
                self._part_type_colors.get(pkey, "")
                or self._part_type_colors.get(str(part_type or "").strip().lower(), "")
                or ""
            ).strip()
            if not raw_color:
                raw_color = "#E8EEF7"
            header_bg = QColor(raw_color) if QColor(raw_color).isValid() else QColor("#7D99B3")
            header_fg = "#FFFFFF" if self._is_dark(header_bg.name()) else "#0F172A"
            row_bg_hex = self._lighten_color(header_bg.name(), 0.68)
            row_border_hex = self._lighten_color(header_bg.name(), 0.52)
            head_row = QWidget()
            head_l = QHBoxLayout(head_row)
            head_l.setContentsMargins(8, 4, 8, 4)
            head_l.setSpacing(8)
            head_cb = QCheckBox()
            head_cb.setChecked(all(bool(self._visibility_map.get(str(r.get("__cutlist_key") or "").strip(), True)) for r in grouped.get(part_type, [])))
            head_cb.toggled.connect(lambda checked, key=pkey: self._set_part_type_visibility(key, checked))
            head_txt = QLabel(part_type)
            head_txt.setStyleSheet(f"QLabel {{ color:{header_fg}; font-size:11px; font-weight:800; background:transparent; border:none; }}")
            head_l.addWidget(head_cb, 0, Qt.AlignmentFlag.AlignVCenter)
            head_l.addWidget(head_txt, 1, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            collapse_btn = QPushButton("+" if pkey in self._collapsed_part_types else "-")
            collapse_btn.setFixedSize(20, 20)
            collapse_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            collapse_btn.setStyleSheet(
                "QPushButton { background:#FFFFFF; color:#334155; border:1px solid #C8D2E0; border-radius:6px; font-size:12px; font-weight:800; padding:0; }"
                "QPushButton:hover { background:#F8FAFC; border-color:#B8C3D4; }"
            )
            collapse_btn.clicked.connect(lambda _=False, key=pkey, btn=collapse_btn: self._toggle_part_type_collapsed(key, btn))
            head_l.addWidget(collapse_btn, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            head_row.setStyleSheet(f"QWidget {{ background:{header_bg.name()}; border:1px solid {header_bg.name()}; border-radius:8px; }}")
            self._visibility_host_l.addWidget(head_row)
            self._visibility_group_widgets.append((head_row, part_type, head_cb, collapse_btn, pkey))
            for row in grouped.get(part_type, []):
                idx = str(row.get("__cutlist_key") or "").strip()
                line = QWidget()
                line_l = QHBoxLayout(line)
                line_l.setContentsMargins(6, 4, 6, 4)
                line_l.setSpacing(8)
                cb = QCheckBox()
                cb.setChecked(bool(self._visibility_map.get(idx, self._boolish(row.get("includeInNesting", True)))))
                cb.toggled.connect(lambda checked, key=idx: self._set_visibility(key, checked))
                name_txt = QLabel(self._row_get(row, "name", "partName") or "Piece")
                name_txt.setWordWrap(True)
                name_txt.setStyleSheet("QLabel { color:#111827; font-size:11px; font-weight:700; background:transparent; border:none; }")
                room_val = self._row_get(row, "room", "section", "jobSection") or "-"
                room_txt = QLabel(room_val)
                room_txt.setWordWrap(True)
                room_txt.setStyleSheet("QLabel { color:#64748B; font-size:10px; background:transparent; border:none; }")
                text_host = QWidget()
                text_host.setStyleSheet("QWidget { background: transparent; border: none; }")
                text_l = QVBoxLayout(text_host)
                text_l.setContentsMargins(0, 0, 0, 0)
                text_l.setSpacing(1)
                text_l.addWidget(name_txt)
                text_l.addWidget(room_txt)
                qty_txt = QLabel(str(self._row_get(row, "quantity", "qty") or "0"))
                qty_txt.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                qty_txt.setStyleSheet("QLabel { color:#111827; font-size:11px; font-weight:700; background:transparent; border:none; min-width:28px; }")
                line_l.addWidget(cb, 0, Qt.AlignmentFlag.AlignTop)
                line_l.addWidget(text_host, 1)
                line_l.addWidget(qty_txt, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                line.setObjectName("cncVisibilityRow")
                line.setStyleSheet(
                    f"QWidget#cncVisibilityRow {{ background:{row_bg_hex}; border:1px solid {row_border_hex}; border-radius:8px; }}"
                    "QWidget#cncVisibilityRow QLabel { background:transparent; border:none; }"
                    "QWidget#cncVisibilityRow QCheckBox { background:transparent; border:none; }"
                    "QWidget#cncVisibilityRow QCheckBox::indicator { width:13px; height:13px; }"
                )
                self._visibility_host_l.addWidget(line)
                search_blob = f"{part_type} {self._row_get(row, 'name', 'partName')} {room_val}".lower()
                self._visibility_row_widgets.append((line, search_blob, part_type.lower(), cb))
        self._visibility_host_l.addStretch(1)
        self._apply_visibility_search()
        self._refresh_group_header_checks()

    def _apply_visibility_search(self) -> None:
        query = str(getattr(self, "_visibility_search", None).text() if isinstance(getattr(self, "_visibility_search", None), QLineEdit) else "").strip().lower()
        visible_groups: set[str] = set()
        for row_widget, blob, group_key, _row_cb in (getattr(self, "_visibility_row_widgets", None) or []):
            collapsed = self._type_key(group_key) in self._collapsed_part_types
            match = ((not query) or (query in blob)) and (not collapsed)
            row_widget.setVisible(match)
            if match:
                visible_groups.add(group_key)
        for group_widget, group_name, _group_cb, collapse_btn, group_key in (getattr(self, "_visibility_group_widgets", None) or []):
            if isinstance(collapse_btn, QPushButton):
                collapse_btn.setText("+" if self._type_key(group_key) in self._collapsed_part_types else "-")
            if not query:
                group_widget.setVisible(True)
            else:
                group_widget.setVisible(group_name.lower() in query or group_name.lower() in visible_groups)

    def _configure_board_table(self, table: QTableWidget) -> int:
        header_fg = "#FFFFFF" if self._is_dark(self._theme_color) else "#0F172A"
        table.setColumnCount(len(self.COLUMNS))
        table.setHorizontalHeaderLabels(self.COLUMNS)
        table.horizontalHeader().setVisible(True)
        table.horizontalHeader().setFixedHeight(22)
        table.verticalHeader().setVisible(False)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        table.setShowGrid(False)
        table.setContentsMargins(0, 0, 0, 0)
        table.setViewportMargins(0, 0, 0, 0)
        table.setStyleSheet(
            "QTableWidget { background:#FFFFFF; alternate-background-color:#F5F5F5; border-left:1px solid #000000; border-right:1px solid #000000; border-top:none; border-bottom:none; border-radius:0px; }"
            "QTableWidget::item { border-right:1px solid #000000; border-bottom:none; }"
            f"QHeaderView {{ background:{self._theme_color}; border:none; margin:0px; padding:0px; }}"
            f"QHeaderView::section {{ background:{self._theme_color}; color:{header_fg}; border:none; border-right:1px solid #000000; padding:2px 6px; font-size:12px; font-weight:800; }}"
            f"QTableCornerButton::section {{ background:{self._theme_color}; border:none; }}"
        )
        table.setColumnWidth(0, 48)
        table.setColumnWidth(1, 118)
        table.setColumnWidth(2, 100)
        table.setColumnWidth(3, 200)
        table.setColumnWidth(4, 260)
        table.setColumnWidth(5, 70)
        table.setColumnWidth(6, 70)
        table.setColumnWidth(7, 70)
        table.setColumnWidth(8, 66)
        table.setColumnWidth(9, 82)
        table.setColumnWidth(10, 82)
        table.setColumnWidth(11, 340)
        if not self._show_grain_column:
            table.setColumnHidden(10, True)
            table.setColumnWidth(11, 420)
        total_w = 0
        for i in range(table.columnCount()):
            if table.isColumnHidden(i):
                continue
            total_w += table.columnWidth(i)
        table.setMinimumWidth(total_w)
        table.setMaximumWidth(total_w)
        return total_w

    def _build_rows(self, rows: list[dict]) -> None:
        normalized: list[dict] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            board_raw = self._row_get(row, "board", "boardType", "material")
            name = self._row_get(row, "name", "partName")
            part_type = self._row_get(row, "partType", "part_type", "type")
            qty = self._row_get(row, "quantity", "qty")
            # Only show actual part rows in CNC table.
            if not name or not part_type or not board_raw or not qty:
                continue
            normalized.append(
                {
                    "__id": self._row_get(row, "__id"),
                    "__cutlist_key": self._row_get(row, "__cutlist_key"),
                    "__raw_row": dict(row),
                    "board_raw": board_raw,
                    "board": self._board_base_label(board_raw),
                    "board_header": self._board_header_label(board_raw),
                    "job_section": self._row_get(row, "room", "section", "jobSection"),
                    "part_type": part_type,
                    "name": name,
                    "height": self._row_get(row, "height", "h"),
                    "width": self._row_get(row, "width", "w"),
                    "depth": self._row_get(row, "depth", "d"),
                    "qty": qty,
                    "clashing": self._row_get(row, "clashing"),
                    "grain": self._row_get(row, "grain"),
                    "info": self._row_get(row, "information", "info", "notes"),
                }
            )

        normalized.sort(key=self._cnc_sort_key)
        if not isinstance(getattr(self, "_boards_host_l", None), QVBoxLayout):
            return
        self._clear_layout(self._boards_host_l)

        standard_rows: list[dict] = []
        cabinetry_rows: list[dict] = []
        for row in normalized:
            if self._is_cabinetry_type(str(row.get("part_type") or "")):
                cabinetry_rows.append(row)
            else:
                standard_rows.append(row)

        grouped: dict[str, list[dict]] = {}
        for row in standard_rows:
            board = str(row.get("board_header") or row.get("board") or "Unknown Board")
            grouped.setdefault(board, []).append(row)

        if not grouped and not cabinetry_rows:
            empty = QLabel("No visible CNC rows.")
            empty.setStyleSheet("QLabel { color:#64748B; font-size:12px; background:transparent; border:none; padding:8px 4px; }")
            self._boards_host_l.addWidget(empty)
            self._boards_host_l.addStretch(1)
            return

        part_counter = 1
        card_row_h = 24
        for board_name in sorted(grouped.keys(), key=lambda x: x.lower()):
            card = QFrame()
            card.setStyleSheet("QFrame { background:transparent; border:none; border-radius:0px; }")
            card_l = QVBoxLayout(card)
            card_l.setContentsMargins(0, 0, 0, 0)
            card_l.setSpacing(0)

            head = QLabel(str(board_name or "Unknown Board"))
            head.setStyleSheet("QLabel { background:#111111; color:#FFFFFF; font-size:12px; font-weight:800; padding:6px 10px; border-top-left-radius:10px; border-top-right-radius:10px; }")
            card_l.addWidget(head)

            table = QTableWidget()
            table_w = self._configure_board_table(table)
            head.setMinimumWidth(table_w)
            head.setMaximumWidth(table_w)
            board_rows = grouped.get(board_name, [])
            table.setRowCount(len(board_rows))
            for row_idx, row in enumerate(board_rows):
                table.setRowHeight(row_idx, card_row_h)
                cells = [
                    str(part_counter),
                    str(row.get("job_section") or ""),
                    str(row.get("part_type") or ""),
                    str(row.get("board") or ""),
                    str(row.get("name") or ""),
                    str(row.get("height") or ""),
                    str(row.get("width") or ""),
                    str(row.get("depth") or ""),
                    str(row.get("qty") or ""),
                    self._format_clashing_display(str(row.get("clashing") or "")),
                    str(row.get("grain") or ""),
                    str(row.get("info") or ""),
                ]
                for col, text in enumerate(cells):
                    align = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
                    if col in (0, 5, 6, 7, 8, 9, 10):
                        align = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignCenter
                    item = self._mk_item(text, align)
                    table.setItem(row_idx, col, item)
                part_type = str(row.get("part_type") or "").strip()
                pkey = self._type_key(part_type)
                pt_color = (
                    self._part_type_colors.get(pkey, "")
                    or self._part_type_colors.get(part_type.lower(), "")
                )
                if pt_color:
                    pt_item = table.item(row_idx, 2)
                    if pt_item is not None:
                        bg = QColor(pt_color)
                        if bg.isValid():
                            fg = QColor("#FFFFFF" if self._is_dark(pt_color) else "#111827")
                            pt_item.setBackground(QBrush(bg))
                            pt_item.setForeground(QBrush(fg))
                            pt_item.setData(Qt.ItemDataRole.BackgroundRole, QBrush(bg))
                            pt_item.setData(Qt.ItemDataRole.ForegroundRole, QBrush(fg))
                            pill = QLabel(part_type)
                            pill.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            pill.setContentsMargins(8, 0, 6, 0)
                            pill.setStyleSheet(
                                "QLabel { "
                                f"background:{bg.name()}; "
                                f"color:{fg.name()}; "
                                "border:none; "
                                "font-size:11px; font-weight:700; "
                                "}"
                            )
                            table.setCellWidget(row_idx, 2, pill)
                part_counter += 1

            table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            total_h = table.horizontalHeader().height()
            for r in range(table.rowCount()):
                total_h += table.rowHeight(r)
            table.setMinimumHeight(total_h)
            table.setMaximumHeight(total_h)
            card_l.addWidget(table)
            bottom_rule = QFrame()
            bottom_rule.setFixedHeight(1)
            bottom_rule.setMinimumWidth(table_w)
            bottom_rule.setMaximumWidth(table_w)
            bottom_rule.setStyleSheet("QFrame { background:#000000; border:none; }")
            card_l.addWidget(bottom_rule)
            self._boards_host_l.addWidget(card)

        if cabinetry_rows:
            cab_card = QFrame()
            cab_card.setStyleSheet("QFrame { background:transparent; border:none; border-radius:0px; }")
            cab_l = QVBoxLayout(cab_card)
            cab_l.setContentsMargins(0, 0, 0, 0)
            cab_l.setSpacing(0)
            cab_head = QLabel("Cabinets")
            cab_head.setStyleSheet("QLabel { background:#111111; color:#FFFFFF; font-size:12px; font-weight:800; padding:6px 10px; border-top-left-radius:10px; border-top-right-radius:10px; }")
            cab_l.addWidget(cab_head)
            grid_host = QWidget()
            grid_host.setStyleSheet("QWidget { background:#FFFFFF; border-left:1px solid #000000; border-right:1px solid #000000; border-bottom:1px solid #000000; }")
            grid = QGridLayout(grid_host)
            grid.setContentsMargins(0, 0, 0, 0)
            grid.setHorizontalSpacing(0)
            grid.setVerticalSpacing(0)
            for col in range(3):
                grid.setColumnStretch(col, 1)

            def _drilling_value(value) -> str:
                txt = str(value or "").strip().lower()
                if not txt:
                    return "No"
                if txt in ("even spacing", "even", "spacing", "equal spacing", "evenly spaced", "even-spaced"):
                    return "Even Spacing"
                if txt in ("centre", "center", "centred", "centered"):
                    return "Centre"
                if txt in ("no", "no drilling", "none", "off", "false", "0"):
                    return "No"
                return str(value or "No")

            left_col_w = 132

            def _cell_row(
                label1: str,
                value1: str,
                stripe: bool = False,
                compact: bool = False,
                label2: str = "",
                value2: str = "",
                value3: str = "",
                shade_left: bool = False,
            ) -> QWidget:
                row_w = QWidget()
                row_w.setFixedHeight(card_row_h)
                row_w.setStyleSheet(f"QWidget {{ background:{'#F5F5F5' if stripe else '#FFFFFF'}; border:none; }}")
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                row_l.setSpacing(0)
                row_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)

                def _add_left_cell(text: str, *, bold: bool = False) -> None:
                    left_cell = QWidget()
                    left_cell.setStyleSheet("QWidget { background:#D9DDE4; border:none; }")
                    left_cell.setFixedWidth(left_col_w)
                    left_cell.setFixedHeight(card_row_h)
                    left_l = QHBoxLayout(left_cell)
                    left_l.setContentsMargins(0, 0, 0, 0)
                    left_l.setSpacing(0)
                    left_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)
                    lb = QLabel(str(text or ""))
                    lb.setStyleSheet(
                        "QLabel { color:#000000; background:transparent; border:none; "
                        f"font-weight:{'700' if bold else '500'}; padding:0px 6px; }}"
                    )
                    left_l.addWidget(lb, 1)
                    row_l.addWidget(left_cell, 0)

                def _mk(text: str, bold: bool = False, stretch: int = 1) -> QLabel:
                    lb = QLabel(str(text or ""))
                    lb.setStyleSheet(
                        "QLabel { color:#000000; background:transparent; border:none; "
                        f"font-weight:{'700' if bold else '500'}; padding:0px 6px; }}"
                    )
                    row_l.addWidget(lb, stretch)
                    return lb

                if compact:
                    # Keep compact rows aligned with the information text start.
                    if shade_left:
                        _add_left_cell(label1, bold=True)
                    else:
                        _mk(label1, True, 20)
                    _mk(value1, False, 1)
                else:
                    if shade_left:
                        _add_left_cell(label1, bold=True)
                    else:
                        _mk(label1, True, 20)
                    _mk(value1, False, 18)
                if label2:
                    _mk(label2, True, 18)
                    _mk(value2, False, 12)
                    _mk(value3, False, 12)
                return row_w

            def _full_row(text: str, *, bold: bool = False, stripe: bool = False, shade_left: bool = False) -> QWidget:
                row_w = QWidget()
                row_w.setFixedHeight(card_row_h)
                row_w.setStyleSheet(f"QWidget {{ background:{'#F5F5F5' if stripe else '#FFFFFF'}; border:none; }}")
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                row_l.setSpacing(0)
                row_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)
                if shade_left:
                    left_pad = QLabel("")
                    left_pad.setStyleSheet("QLabel { background:#D9DDE4; border:none; }")
                    left_pad.setFixedWidth(left_col_w)
                    left_pad.setFixedHeight(card_row_h)
                    row_l.addWidget(left_pad, 0)
                lb = QLabel(str(text or ""))
                lb.setStyleSheet(
                    "QLabel { color:#000000; background:transparent; border:none; "
                    f"font-weight:{'700' if bold else '500'}; padding:0px 6px; }}"
                )
                row_l.addWidget(lb, 1)
                return row_w

            def _info_value_row(text: str, *, stripe: bool = False) -> QWidget:
                # Align info text start with the shelf quantity value column.
                row_w = QWidget()
                row_w.setFixedHeight(card_row_h)
                row_w.setStyleSheet(f"QWidget {{ background:{'#F5F5F5' if stripe else '#FFFFFF'}; border:none; }}")
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                row_l.setSpacing(0)
                row_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)

                pad = QLabel("")
                pad.setStyleSheet("QLabel { background:#D9DDE4; border:none; }")
                pad.setFixedWidth(left_col_w)
                pad.setFixedHeight(card_row_h)
                row_l.addWidget(pad, 0)

                lb = QLabel(str(text or ""))
                lb.setStyleSheet(
                    "QLabel { color:#000000; background:transparent; border:none; "
                    "font-weight:500; padding:0px 6px; }"
                )
                row_l.addWidget(lb, 60)
                return row_w

            def _info_header_row(first_text: str, *, stripe: bool = False) -> QWidget:
                # Keep label on the left while first info line starts at the value column.
                row_w = QWidget()
                row_w.setFixedHeight(card_row_h)
                row_w.setStyleSheet(f"QWidget {{ background:{'#F5F5F5' if stripe else '#FFFFFF'}; border:none; }}")
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                row_l.setSpacing(0)
                row_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)

                left_cell = QWidget()
                left_cell.setStyleSheet("QWidget { background:#D9DDE4; border:none; }")
                left_cell.setFixedWidth(left_col_w)
                left_cell.setFixedHeight(card_row_h)
                left_l = QHBoxLayout(left_cell)
                left_l.setContentsMargins(0, 0, 0, 0)
                left_l.setSpacing(0)
                left_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)
                label = QLabel("Information")
                label.setStyleSheet(
                    "QLabel { color:#000000; background:transparent; border:none; "
                    "font-weight:700; padding:0px 6px; }"
                )
                left_l.addWidget(label, 1)
                row_l.addWidget(left_cell, 0)

                val = QLabel(str(first_text or "-"))
                val.setStyleSheet(
                    "QLabel { color:#000000; background:transparent; border:none; "
                    "font-weight:500; padding:0px 6px; }"
                )
                row_l.addWidget(val, 1)
                return row_w

            has_any_cabinet_info = False
            for _r in cabinetry_rows:
                _src = self._find_source_row_for(_r) or {}
                _info_raw = str(
                    self._first_nonempty(
                        _r.get("info"),
                        self._row_get(_src, "information", "info", "notes"),
                    )
                    or ""
                ).strip()
                if _info_raw:
                    has_any_cabinet_info = True
                    break

            def _id_name_row(cab_id_txt: str, cab_name_txt: str) -> QWidget:
                row_w = QWidget()
                row_w.setFixedHeight(card_row_h)
                id_bg = self._normalize_color(self._theme_color)
                id_fg = "#FFFFFF" if self._is_dark(id_bg) else "#0F172A"
                row_w.setStyleSheet(f"QWidget {{ background:{id_bg}; border:none; }}")
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                row_l.setSpacing(2)
                row_l.setAlignment(Qt.AlignmentFlag.AlignVCenter)

                def _lbl(text: str, bold: bool = False) -> QLabel:
                    lb = QLabel(str(text or ""))
                    lb.setStyleSheet(
                        f"QLabel {{ color:{id_fg}; background:transparent; border:none; "
                        f"font-weight:{'700' if bold else '500'}; padding:0px 6px; }}"
                    )
                    row_l.addWidget(lb, 0)
                    return lb

                _lbl("ID:", True)
                _lbl(str(cab_id_txt or "-"), False)
                row_l.addSpacing(10)
                _lbl("Name:", True)
                name_lbl = _lbl(str(cab_name_txt or "-"), False)
                name_lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
                row_l.addStretch(1)
                return row_w

            for i, row in enumerate(cabinetry_rows):
                cab_id = str(part_counter)
                part_counter += 1

                source_row = self._find_source_row_for(row) or {}
                name = str(row.get("name") or "-")
                material = str(row.get("board") or "-")
                info_raw = str(
                    self._first_nonempty(
                        row.get("info"),
                        self._row_get(source_row, "information", "info", "notes"),
                    )
                    or ""
                )
                info_lines = [ln.strip() for ln in info_raw.replace("\r", "\n").split("\n") if ln.strip()]
                h = str(row.get("height") or "-")
                w = str(row.get("width") or "-")
                d = str(row.get("depth") or "-")
                qty_txt = str(row.get("qty") or "-")
                size_txt = f"H:{h} x W:{w} x D:{d}"
                board_key = str(
                    self._first_nonempty(
                        self._row_get(source_row, "board", "boardType", "material"),
                        self._row_get(row, "board_raw", "board"),
                    )
                ).strip()
                edging_txt = str(self._board_edging_map.get(board_key) or "").strip()
                clashing_txt = edging_txt or "-"

                shelves_summary_row = self._parse_cabinet_shelves_summary(self._row_get(row, "clashing", "shelves"))
                shelves_summary_src = self._parse_cabinet_shelves_summary(self._row_get(source_row, "clashing", "shelves"))
                adj_num = self._first_nonempty(
                    shelves_summary_src.get("adjustable_qty"),
                    shelves_summary_row.get("adjustable_qty"),
                    self._row_get(
                        source_row,
                        "adjustable_shelves",
                        "adjustableShelf",
                        "adjustableShelves",
                        "adjustable_shelf",
                        "adjustable",
                        "adjustableCount",
                        "adjustable_count",
                    ),
                    self._row_get(
                        row,
                        "adjustable_shelves",
                        "adjustableShelf",
                        "adjustableShelves",
                        "adjustable_shelf",
                        "adjustable",
                        "adjustableCount",
                        "adjustable_count",
                    ),
                ) or "-"
                adj_drill = _drilling_value(
                    self._first_nonempty(
                        shelves_summary_src.get("adjustable_drilling"),
                        shelves_summary_row.get("adjustable_drilling"),
                        self._row_get(
                            source_row,
                            "adjustable_drilling",
                            "adjustableShelfDrilling",
                            "adjustableDrilling",
                            "adjustable_shelf_drilling",
                            "adjustable_drill",
                        ),
                        self._row_get(
                            row,
                            "adjustable_drilling",
                            "adjustableShelfDrilling",
                            "adjustableDrilling",
                            "adjustable_shelf_drilling",
                            "adjustable_drill",
                        ),
                    )
                )
                adj_pos = ""

                fix_num = self._first_nonempty(
                    shelves_summary_src.get("fixed_qty"),
                    shelves_summary_row.get("fixed_qty"),
                    self._row_get(
                        source_row,
                        "fixed_shelves",
                        "fixedShelf",
                        "fixedShelves",
                        "fixed_shelf",
                        "fixed",
                        "fixedCount",
                        "fixed_count",
                    ),
                    self._row_get(
                        row,
                        "fixed_shelves",
                        "fixedShelf",
                        "fixedShelves",
                        "fixed_shelf",
                        "fixed",
                        "fixedCount",
                        "fixed_count",
                    ),
                ) or "-"
                fix_drill = _drilling_value(
                    self._first_nonempty(
                        shelves_summary_src.get("fixed_drilling"),
                        shelves_summary_row.get("fixed_drilling"),
                        self._row_get(
                            source_row,
                            "fixed_drilling",
                            "fixedShelfDrilling",
                            "fixedDrilling",
                            "fixed_shelf_drilling",
                            "fixed_drill",
                        ),
                        self._row_get(
                            row,
                            "fixed_drilling",
                            "fixedShelfDrilling",
                            "fixedDrilling",
                            "fixed_shelf_drilling",
                            "fixed_drill",
                        ),
                    )
                )
                fix_pos = ""

                box = QFrame()
                box.setStyleSheet("QFrame { background:#FFFFFF; border:none; border-left:1px solid #000000; border-right:1px solid #000000; border-bottom:1px solid #000000; border-radius:0px; }")
                box_l = QVBoxLayout(box)
                box_l.setContentsMargins(0, 0, 0, 0)
                box_l.setSpacing(0)

                box_l.addWidget(_id_name_row(cab_id, name))
                box_l.addWidget(_cell_row("Cabinet Size", size_txt, True, True, shade_left=True))
                box_l.addWidget(_cell_row("Quantity", qty_txt, False, True, shade_left=True))
                box_l.addWidget(_cell_row("Material", material, True, True, shade_left=True))
                box_l.addWidget(_cell_row("Clashing", clashing_txt, False, True, shade_left=True))
                box_l.addWidget(_full_row("", stripe=True, shade_left=True))
                box_l.addWidget(_cell_row("Adjustable Shelf", adj_num, False, False, "Drilling:", adj_drill, adj_pos, shade_left=True))
                box_l.addWidget(_cell_row("Fixed Shelf", fix_num, True, False, "Drilling:", fix_drill, fix_pos, shade_left=True))
                if has_any_cabinet_info:
                    first_line = f"- {info_lines[0]}" if info_lines else "-"
                    box_l.addWidget(_info_header_row(first_line, stripe=False))
                    for idx_ln, ln in enumerate(info_lines[1:]):
                        box_l.addWidget(_info_value_row(f"- {ln}", stripe=(idx_ln % 2 == 0)))
                else:
                    box_l.addWidget(_full_row("", stripe=False, shade_left=True))

                grid.addWidget(box, i // 3, i % 3)

            cab_l.addWidget(grid_host)
            self._boards_host_l.addWidget(cab_card)

        self._boards_host_l.addStretch(1)
